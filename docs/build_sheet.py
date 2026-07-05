#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

seen = pickle.load(open('.firecrawl/seen.pkl', 'rb'))
items = list(seen.items())
def U(i): return items[i][0]

# Annahar publish dates (from page metadata, scraped)
annahar_dates = {
 0:'2026-04-13', 6:'2023-09-19', 24:'2026-05-13', 32:'2023-08-06', 33:'2026-05-29',
 34:'2024-09-19', 35:'2023-08-31', 36:'2024-04-16', 37:'2024-01-04', 38:'2023-12-07',
 39:'2023-10-22', 40:'2021-03-01', 41:'2026-02-14', 42:'2022-06-09', 43:'2021-05-27',
 44:'2021-10-29', 62:'2025-03-11',
}

# ARTICLES authored by Dr. Wadiha El Amiouni  -> (index, clean_title, source, date, topic)
articles = [
 (0,  'الحروب وموازين القوى في زمن الذكاء الاصطناعي', 'النهار / Annahar', annahar_dates[0], 'الذكاء الاصطناعي / الحروب'),
 (41, "سيكولوجيا «الجزيرة المحرمة» لشبكة إبستين وديناميكيات السلطة المطلقة", 'النهار / Annahar', annahar_dates[41], 'علم النفس / السلطة'),
 (24, 'أزمة لبنان: الميثاقية والدستور بين التعطيل والإصلاح', 'النهار / Annahar', annahar_dates[24], 'الدستور / السياسة'),
 (33, 'مئوية الدستور اللبناني بين دولة المؤسسات وميثاقية التعطيل', 'النهار / Annahar', annahar_dates[33], 'الدستور / السياسة'),
 (34, 'التكنولوجيا والحروب: تأثير متزايد في الصراع', 'النهار / Annahar', annahar_dates[34], 'التكنولوجيا / الحروب'),
 (37, "«تكنولوجيا الحزب»: بناء هياكل سياسية قوية في عصر الرقمنة", 'النهار / Annahar', annahar_dates[37], 'الرقمنة / السياسة'),
 (36, 'العملات الرقمية بين الاستخدام والتنظيم', 'النهار / Annahar', annahar_dates[36], 'العملات الرقمية'),
 (6,  'الذكاء الاصطناعي والتحديات القانونية في لبنان', 'النهار / Annahar', annahar_dates[6], 'الذكاء الاصطناعي / القانون'),
 (35, 'القمار الإلكتروني في لبنان: تحديات اجتماعية وقانونية', 'النهار / Annahar', annahar_dates[35], 'القمار الإلكتروني'),
 (39, 'التجارة الرقمية وقانون التوقيع الإلكتروني: مفتاح النهضة الاقتصادية في لبنان', 'النهار / Annahar', annahar_dates[39], 'الاقتصاد الرقمي'),
 (38, 'المنصة الرقمية في مرفأ بيروت: استعادة لدوره الاقتصادي', 'النهار / Annahar', annahar_dates[38], 'الاقتصاد الرقمي'),
 (32, 'استراتيجية التحول الرقمي والقوانين التشريعية اللازمة في لبنان', 'النهار / Annahar', annahar_dates[32], 'التحول الرقمي'),
 (42, 'الحكومة الإلكترونية والحوكمة الرشيدة بين الإصلاح والفساد في لبنان', 'النهار / Annahar', annahar_dates[42], 'الحوكمة الرقمية'),
 (44, 'الجريمة الرقمية كيف نكافحها؟', 'النهار / Annahar', annahar_dates[44], 'الجريمة الرقمية'),
 (43, "إطلاق حملة «دخانك بيقتلني»... توصية بتفعيل قانون الحد من التدخين", 'النهار / Annahar', annahar_dates[43], 'الصحة العامة'),
 (40, 'الوعي السياسي المتخلّف', 'النهار / Annahar', annahar_dates[40], 'السياسة / المجتمع'),
 (62, 'الحكواتي السياسي: بين فنّ السرد وعجز الفعل', 'النهار / Annahar', annahar_dates[62], 'السياسة / المجتمع'),
 (63, 'الانتخابات البلدية: إصلاح الإدارة المحلية مدخل إلى الدولة الحديثة', 'النهار (English ed.) / Annahar', '', 'الانتخابات البلدية'),
 # Safir al-Shamal column (date from URL)
 (78, 'مدينة صور تحت النار', 'سفير الشمال / Safir al-Shamal', '2026-05-29', 'سياسة / حرب'),
 (69, 'عودة الميت إلى الحياة في زمن الذكاء الاصطناعي', 'سفير الشمال / Safir al-Shamal', '2025-12-15', 'الذكاء الاصطناعي'),
 (64, 'الحكومات السرية: هل تتحكم قوى خفية بمصير لبنان؟', 'سفير الشمال / Safir al-Shamal', '2025-03-30', 'سياسة'),
 (50, 'العملات الرقمية بين الاستخدام والتنظيم', 'سفير الشمال / Safir al-Shamal', '2024-04-16', 'العملات الرقمية'),
 (12, 'الذكاء الاصطناعي بين الأخطار والمزايا والقوانين اللازمة في لبنان', 'سفير الشمال / Safir al-Shamal', '2023-08-21', 'الذكاء الاصطناعي'),
 # INN Lebanon
 (73, '«ممر داوود»: عزل اقتصادي وتكنولوجي يهدد ما تبقّى من لبنان', 'INN Lebanon', '', 'سياسة / اقتصاد'),
 (22, 'الحكومات السرية: هل تتحكم قوى خفية بمصير لبنان؟', 'INN Lebanon', '', 'سياسة'),
 (19, 'الانتخابات البلدية: إصلاح الإدارة المحلية مدخل إلى الدولة الحديثة', 'INN Lebanon', '', 'الانتخابات البلدية'),
 (68, 'الإعلام الرقمي في لبنان بين حرية التعبير وضرورة التنظيم', 'INN Lebanon', '', 'الإعلام الرقمي'),
 # Other republishers / outlets
 (76, 'الذكاء الاصطناعي التركيبي: الآلة تتجاوز الإنسان ولبنان ما زال في نقطة الصفر', 'مرصد نيوز / Marsad News', '', 'الذكاء الاصطناعي'),
 (20, 'العملات الرقمية (Crypto Currencies) بين الإيجابيات والسلبيات', 'مجلة عرب أستراليا / Arabs Australia', '', 'العملات الرقمية'),
 # Folded in from Facebook cross-check (canonical, non-FB link)
 ('https://beirut2030.me/?p=5041', 'كيف يفهم الأطفال الموت... وكيف ندعمهم في مواجهة فقدان القريب؟ (تصريح)', 'Beirut 2030', '2026-03-26', 'علم النفس / الطفولة'),
 # Sourced from her CV (تموز 2026) — verified byline matches on مجلة الحداثة archive
 ('https://alhadathamagazine.blogspot.com/2016/12/blog-post.html', 'التعليم والتكنولوجيا في المؤسسات التربوية: استخدام التقنيات الحديثة في المدارس الخاصة', 'مجلة الحداثة / Al-Hadatha Magazine', '2016-10-01', 'التربية والتكنولوجيا'),
 ('https://alhadathamagazine.blogspot.com/2016/01/blog-post_26.html', 'الإشكاليات النسوية ودورها في تأسيس علم اجتماع المرأة والإشكاليات الميدانية', 'مجلة الحداثة / Al-Hadatha Magazine', '2016-01-01', 'المرأة / علم الاجتماع'),
 ('https://alhadathamagazine.blogspot.com/2017/03/blog-post_13.html', 'التكاذب الافتراضي على مواقع التواصل الاجتماعي', 'مجلة الحداثة / Al-Hadatha Magazine', '2017-04-01', 'المجتمع الرقمي'),
 ('https://alhadathamagazine.blogspot.com/2014/01/blog-post.html', 'المرأة العربية وجرائم الشرف', 'مجلة الحداثة / Al-Hadatha Magazine', '2014-04-01', 'المرأة'),
 ('https://alhadathamagazine.blogspot.com/2012/01/blog-post.html', 'العنف في الخطاب السياسي العربي', 'مجلة الحداثة / Al-Hadatha Magazine', '2012-07-01', 'السياسة / العنف'),
 ('https://alhadathamagazine.blogspot.com/2008/01/blog-post.html', 'المرأة في عصر النهضة', 'مجلة الحداثة / Al-Hadatha Magazine', '2008-10-01', 'المرأة / التاريخ'),
]

# MEDIA appearances -> (index, clean_title, program/source, date, topic)
media = [
 (140,'برنامج Leadership – المعنى الوطني للانتخابات','MTV (MTV Alive)','2025-12-03','انتخابات'),
 (55, 'برنامج Leadership – المواطنة والعمل البلدي','MTV (MTV Alive) / YouTube','2025-09-17','المواطنة'),
 (134,'برنامج Leadership – معنى الاستقلال في الحرب','MTV (MTV Alive)','2024-11-22','الاستقلال'),
 (142,'برنامج Leadership – القيادة والحروب','MTV (MTV Alive)','2024-04-17','القيادة'),
 (9,  'برنامج Leadership – القيادة والابتكار في عالم متغيّر','MTV / YouTube','2024-02-07','القيادة'),
 (156,'برنامج Leadership – القيادة في الحركات الكشفية','MTV / YouTube','2023-01-30','القيادة'),
 (153,'برنامج Leadership – مفاعيل القيادة السيّئة','MTV / YouTube','2022-09-19','القيادة'),
 (152,'برنامج Leadership – القيادة والتحوّل الرقمي','MTV / YouTube','2022-07-25','القيادة'),
 (150,'برنامج Leadership – القيادة والسياسة','MTV / YouTube','2022-04-04','القيادة'),
 (135,'برنامج Leadership – القيادة والحروب (VOD)','MTV (mtv.com.lb)','','القيادة'),
 (136,'برنامج Family – كيف أتعلّم حب الوطن','MTV / YouTube','2021-11-22','التربية الوطنية'),
 (157,'برنامج Family – التفاعل الفوقي في المجتمع','MTV / YouTube','2020-11-10','المجتمع'),
 (146,'برنامج Family – التربية على المواطنة','MTV / YouTube','','المواطنة'),
 (116,'برنامج Family – الإدمان الرقمي بزمن الكورونا (VOD)','MTV (mtv.com.lb)','','الإدمان الرقمي'),
 (117,'برنامج «علمي علمك» – ضيفة الحلقة','إذاعة لبنان الحر / RLL','2026-01-12','علوم اجتماعية'),
 (124,'ضيفة تلفزيون لبنان مع بيا مخول','تلفزيون لبنان / Télé Liban','','مقابلة عامة'),
 (104,'باحثة اجتماعية وأستاذة جامعية – مقابلة','تلفزيون لبنان / Télé Liban','','مقابلة عامة'),
 (149,'برنامج «نقطة فاصلة» – مجتمع بين الواقع والافتراض','OTV Lebanon','','المجتمع الرقمي'),
 (115,'برنامج «باحثات» – الحلقة 32','تلفزيون مريم / Mariam TV','','علوم اجتماعية'),
 (112,"برنامج «315» – ضيفة الحلقة",'لبنان ديبايت / RED TV','','مقابلة'),
 (132,'مقابلة لـ RED TV: الأزمة الاقتصادية تؤخر الزواج والإنجاب','لبنان ديبايت / RED TV','','الزواج / الاقتصاد'),
 (121,'برنامج «أحلى صباح» – مقابلة','HMG / تلفزيون لبنان','','مقابلة صباحية'),
 (120,'برنامج «مشوار امرأة» – ضيفة الحلقة','YouTube / Facebook','','مقابلة'),
 (133,'برنامج «بتحلى الحياة» – الكاتبة وديعة الأميوني','YouTube','','مقابلة'),
 (118,'برنامج «طبيب أونلاين» – حلقة اجتماعية نفسية','Instagram','','الصحة النفسية'),
 (161,'برنامج «شمس الصباح» – حماية المعلومات الشخصية على الإنترنت','YouTube','','الأمن الرقمي'),
 (2,  'تكشف تأثير الذكاء الاصطناعي على المجتمع','YouTube','','الذكاء الاصطناعي'),
 (147,'الذكاء اللبناني بين التفوّق الفردي وفشل الدولة','YouTube','','المجتمع'),
 (148,"لعبة Labubu و«جنون التقليد»... ماذا يقول علم النفس؟",'الجديد / Al Jadeed','2025-06-18','علم النفس'),
 (160,'عصر مركّب بزمن الاستقلالية: كيف تغيّر مفهوم العنوسة','YouTube','','المجتمع / المرأة'),
 (110,'نسبة البطالة وصلت اليوم في لبنان...','YouTube','','البطالة'),
 (46, 'لبنان يعيش فترة مأساوية على جميع المستويات بسبب العدوان','YouTube','','الوضع اللبناني'),
 (145,'أهمية مبادرات المجتمع المدني','YouTube','','المجتمع المدني'),
 (154,'المواطن اللبناني يعيش أزمات سياسية واقتصادية متلاحقة','YouTube','','الوضع اللبناني'),
 (155,'هل مواجهة المشاكل بالفكاهة حل؟','YouTube','','علم النفس'),
 (159,'إلى أين تتجه المواجهة بين الحكومات ومنصات التواصل؟','YouTube','','الإعلام الرقمي'),
 (138,'برنامج «لبنان اليوم»','YouTube','2020-11-21','مقابلة عامة'),
 (128,'القاهرة الإخبارية – مقابلة (أين الأمم المتحدة؟)','القاهرة الإخبارية / Nabd','','شؤون دولية'),
 # Folded in from Facebook cross-check (canonical YouTube link)
 ('https://www.youtube.com/watch?v=qHYFkEeiFx4','برنامج Family – دور التربية في تفادي الفتن المجتمعية','MTV / YouTube','','التربية / المجتمع'),
]

# Build workbook
wb = Workbook()
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill('solid', fgColor='1F4E78')
hdr_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=14, color='1F4E78')
wrap = Alignment(wrap_text=True, vertical='top', horizontal='right', readingOrder=2)
wrapL = Alignment(wrap_text=True, vertical='top', horizontal='left')

def make_sheet(ws, heading, rows):
    ws.sheet_view.rightToLeft = True
    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = heading; c.font = title_font; c.alignment = Alignment(horizontal='right')
    headers = ['#', 'العنوان / Title', 'المصدر / Source', 'التاريخ / Date', 'الرابط / Link']
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    r = 3
    for n, (idx, title, src, date, topic) in enumerate(rows, 1):
        url = U(idx) if isinstance(idx, int) else idx
        ws.cell(row=r, column=1, value=n).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=r, column=2, value=title).alignment = wrap
        ws.cell(row=r, column=3, value=src).alignment = wrap
        ws.cell(row=r, column=4, value=date or '—').alignment = Alignment(horizontal='center', vertical='top')
        lc = ws.cell(row=r, column=5, value=url)
        lc.hyperlink = url; lc.font = Font(color='0563C1', underline='single'); lc.alignment = wrapL
        for j in range(1, 6):
            ws.cell(row=r, column=j).border = border
        r += 1
    widths = [5, 52, 26, 13, 60]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 26

# NEW items surfaced by cross-checking her Facebook (https://www.facebook.com/namiouni)
# -> (title, source/program, date, link, status)
fb_new = [
 ('برنامج Family – دور التربية في تفادي الفتن المجتمعية', 'MTV / YouTube', '', 'https://www.youtube.com/watch?v=qHYFkEeiFx4', 'NEW — MTV Family episode not in original list'),
 ('مؤشر السعادة العالمي – لماذا نجحت دول في إسعاد شعوبها (هاتفياً)', 'Mosaïque FM (تونس) – بثّ مباشر', '2026-06-21', 'https://www.facebook.com/namiouni', 'NEW outlet — Tunisian radio, recurring guest'),
 ('مؤشر السعادة العالمي (حلقة ثانية)', 'Mosaïque FM (تونس) – بثّ مباشر', '2026-06-14', 'https://www.facebook.com/namiouni', 'NEW outlet'),
 ('الألعاب الإلكترونية وحماية الأطفال / الابتزاز الإلكتروني', 'Mosaïque FM (تونس)', '2026-02-15', 'https://www.facebook.com/namiouni', 'NEW outlet'),
 ('مقابلة: علم الاجتماع، الاغتراب، التنمر', 'Catchy Talk', '2026-06', 'https://www.facebook.com/namiouni', 'NEW program'),
 ('كيف يفهم الأطفال الموت... وكيف ندعمهم في مواجهة فقدان القريب؟ (مقابلة/تصريح)', 'Beirut 2030', '2026-03-26', 'https://beirut2030.me/?p=5041', 'NEW platform — she is quoted/interviewed'),
 ('برنامج «طبيب أونلاين» – حلقة إنسانية/اجتماعية/نفسية', 'إذاعة لبنان / Radio Liban', '2026-03-02', 'https://www.facebook.com/RadioLibanPage/videos/1029887119381165/', 'Refines existing entry — confirms Radio Liban host'),
]

def make_fb_sheet(ws, heading, rows):
    ws.sheet_view.rightToLeft = True
    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = heading; c.font = title_font; c.alignment = Alignment(horizontal='right')
    headers = ['#', 'العنوان / Title', 'المصدر / Source', 'التاريخ / Date', 'الرابط / Link', 'ملاحظة / Status']
    ws.merge_cells('A1:F1')
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    r = 3
    for n, (title, src, date, url, status) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=n).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=r, column=2, value=title).alignment = wrap
        ws.cell(row=r, column=3, value=src).alignment = wrap
        ws.cell(row=r, column=4, value=date or '—').alignment = Alignment(horizontal='center', vertical='top')
        lc = ws.cell(row=r, column=5, value=url)
        lc.hyperlink = url; lc.font = Font(color='0563C1', underline='single'); lc.alignment = wrapL
        ws.cell(row=r, column=6, value=status).alignment = wrap
        for j in range(1, 7):
            ws.cell(row=r, column=j).border = border
        r += 1
    for j, w in enumerate([5, 46, 26, 13, 50, 34], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 26

ws1 = wb.active; ws1.title = 'Articles مقالات'
make_sheet(ws1, 'مقالات بقلم د. وديعة الأميوني — Articles by Dr. Wadiha El Amiouni', articles)
ws2 = wb.create_sheet('Media ظهور إعلامي')
make_sheet(ws2, 'الظهور الإعلامي لـ د. وديعة الأميوني — Media appearances', media)

# Notes/about sheet
ws3 = wb.create_sheet('About')
ws3.sheet_view.rightToLeft = True
notes = [
 'د. وديعة الأميوني — Dr. Wadiha El Amiouni',
 'أستاذة (بروفيسور) في علم الاجتماع، الجامعة اللبنانية — مديرة معهد العلوم الاجتماعية (الفرع الثالث).',
 'الملف الجامعي: https://ul.edu.lb/en/teachers/wadiha-el-amiouni/5941',
 '',
 'ملاحظات منهجية / Notes:',
 '• تم جمع الروابط عبر بحث ثنائي اللغة (عربي/إنجليزي) على الويب (Firecrawl) في حزيران 2026.',
 '• «Articles» = مقالات بقلمها. «Media» = ظهور تلفزيوني/إذاعي/فيديو.',
 '• تواريخ مقالات النهار مأخوذة من بيانات الصفحة (datePublished). تواريخ سفير الشمال من رابط المقال.',
 '• تواريخ بعض الفيديوهات مستخرجة من عنوان الحلقة (DD/MM/YYYY)؛ الخلية «—» تعني أن التاريخ غير مثبت في العنوان (متوفر على صفحة الرابط).',
 '• تمّت تصفية شخص آخر يحمل اسماً مشابهاً (وديعة ابراهيم الأميوني «نبال»، أستاذ ألسنية مواليد 1945) ومقاطع لضيوف آخرين.',
 '• عدة مقالات أعيد نشرها في أكثر من موقع (النهار، سفير الشمال، INN، مرصد نيوز...)؛ أُدرج المصدر الأساسي مع نسخ بديلة عند الحاجة.',
 '',
 '• جرت مراجعة صفحتها على فيسبوك (namiouni) للتأكد من التغطية؛ العناصر ذات الروابط الأصلية (خارج فيسبوك) أُدرجت ضمن قسمي «مقالات» و«ظهور إعلامي» — مثل حلقة Family (يوتيوب) وتصريح Beirut 2030.',
 '• عناصر تظهر حصراً على فيسبوك بلا رابط مصدر خارجي (بثوث موزاييك آف آم المباشرة، Catchy Talk) لم تُدرَج، تفادياً لروابط فيسبوك.',
]
for i, t in enumerate(notes, 1):
    cell = ws3.cell(row=i, column=1, value=t)
    cell.alignment = Alignment(horizontal='right', wrap_text=True)
    if i == 1: cell.font = title_font
ws3.column_dimensions['A'].width = 110

out = 'Wadiha_El_Amiouni_Articles_and_Media.xlsx'
wb.save(out)
print('SAVED', out, '| articles:', len(articles), '| media:', len(media))
